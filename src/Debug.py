import os
import cv2
import json
from datetime import datetime

from openpyxl import Workbook, load_workbook

with open('settings.json') as f:
	setData = json.load(f)
	misc = setData['Misc']
	trouble = misc['Trouble']
	colors = setData['colours']

def cvWin(img):
	cv2.namedWindow("img",cv2.WINDOW_FREERATIO)
	cv2.imshow("img",img)
	cv2.waitKey(0)

def drawCNT(img,cnt):
	cntImg = cv2.drawContours(img,[cnt],0,(255,255,255),1)
	cv2.namedWindow("img",cv2.WINDOW_FREERATIO)
	cv2.imshow("img",cntImg)
	cv2.waitKey(0)

def saveImg(img,loc,timestp):
	if not trouble:
		imgdir = os.path.join(os.path.dirname(os.path.dirname(__file__)),loc,datetime.today().strftime("%b%y"))
		if not os.path.exists(imgdir): os.makedirs(imgdir)
		imgfile = os.path.join(imgdir,timestp+".png")
		if not os.path.exists(imgfile):
			cv2.imwrite(imgfile,img)

def dataLog(name,log):
	fileName = name + "_" + datetime.today().strftime("%d-%m-%y") if name[:3] == "Acc" else name
	foldir = os.path.join(os.path.dirname(os.path.dirname(__file__)),"log",datetime.today().strftime("%b%y"))
	if not os.path.exists(foldir): os.makedirs(foldir)
	filepath = os.path.join(foldir,fileName+".xlsx")

	header = ["Time","Block Name","Cali Pin","Color","PixelArea","Area","SilverRng","GoldRng","RoseGoldRng","BlackRng"]

	log.insert(1,name)
	for i in colors:
		log.append(str(list(colors[i].values())))

	if os.path.exists(filepath):
		wb = load_workbook(filepath)
		newRow = wb.active.max_row + 1
		
		for i, data in enumerate(log):
			wb.active.cell(row=newRow,column=i+1).value = data

	else:
		wb = Workbook()

		for i, data in enumerate(log):
			wb.active.cell(row=1,column=i+1).value = header[i]
			wb.active.cell(row=2,column=i+1).value = data

	wb.save(filepath)
	