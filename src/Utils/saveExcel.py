import os
from openpyxl import *

class Excel():
    def __init__(self,filePath,defVar,input=False):
        self.saveExcel(filePath,defVar,input)

    def saveExcel(self,filePath,defVar,input):
        if os.path.exists(filePath):
            wb = load_workbook(filePath)
            maxCol = wb.active.max_column
            newCol = maxCol + 1
            if input:
                for r in wb.active.iter_rows(max_row=wb.active.max_row,max_col=1):
                    for c in r:
                        if c.value == input:
                            if maxCol == 1: maxCol = newCol
                            wb.active.cell(row=c.row,column=maxCol).value = defVar[input].cget("text")
            else:
                for i, defs in enumerate(defVar): 
                    wb.active.cell(row=i+1,column=newCol).value = defVar[defs].cget("text")
        else:
            wb = Workbook()
            for i, defs in enumerate(defVar):
                wb.active.cell(row=i+1,column=1).value = defs
                wb.active.cell(row=i+1,column=2).value = defVar[defs].cget("text")

        wb.save(filePath)