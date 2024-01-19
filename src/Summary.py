import os
from tkinter import *
from openpyxl import *
from tkinter.ttk import Separator

from Utils.readSettings import readSettings
from Utils.prass import PRASS

class Summary(readSettings):
    def __init__(self,root,inData,filePath,Wscreen,Hscreen):
        super().__init__()
        self.initialize(inData,filePath,Wscreen,Hscreen)
        self.reset(root)

    def reset(self,root):
        self.root = Toplevel(root)
        self.win_config()
        self.widgets()
        self.load_SumData(root)
        self.root.grab_set()
        self.root.mainloop()

    def initialize(self,inData,filePath,Wscreen,Hscreen):
        self.res = True
        self.filePath = filePath
        self.inData = inData
        self.Wscreen,self.Hscreen = Wscreen,Hscreen

    def win_config(self):
        self.root.title("Summary Window")
        self.frame = Frame(self.root)
        self.frame.pack()

    def widgets(self):
        # Labelframe for Lot Number Data
        ####################################################################################################

        lotNumCont = LabelFrame(self.frame, bd=5, relief=FLAT)
        lotNumCont.grid(row=0, column=0, padx=3, pady=1,sticky=W)

        lotNumTxt = Label(lotNumCont, text="Lot Number: ", font=self.font['S'])
        lotNumTxt.grid(row=0, column=0, padx=3, pady=1, sticky=W)

        lotNum = Label(lotNumCont, text=self.inData[0], font=self.font['S'])
        lotNum.grid(row=0, column=1, padx=3, pady=1)

        # Labelframe for Defects Data
        ####################################################################################################

        self.dataCont = LabelFrame(self.frame, bd=5, relief=FLAT)
        self.dataCont.grid(row=1, column=0, pady=1, padx=10)

        # Block No 
        blockTxt = Label(self.dataCont, font=self.font['S'], text="Block No")
        blockTxt.grid(row=0, column=0, pady=1, padx=15, sticky=W)

    ####################################################################################################
    """
    Functions Used for the Widgets Above
    """ 
    ####################################################################################################

    def load_SumData(self,root):
        if os.path.exists(self.filePath):
            wb = load_workbook(self.filePath)
            ws = wb.active

            maxcol = ws.max_column
            maxrow = ws.max_row

            xVal = 300 if maxcol*90 < 300 else maxcol*90
            if self.Wscreen < xVal*1.1: self.root.state('zoomed')
            else: self.root.geometry(str(xVal)+f"x{int(self.Hscreen*0.85)}+0+0")
            
            Separator(self.frame, orient="horizontal").grid(row=0, column=0, columnspan=maxcol, sticky=EW+S)

            for i in range(maxcol):
                if i != 0:             
                    Button(self.dataCont, text="x", font=self.font['S'], height=1, width=2, bg="#fa6464", command= lambda i=i: self.delete(ws, wb, i,root)).grid(row=maxrow+1,column=i,pady=5,padx=15,sticky=S)
                    Label(self.dataCont, font=self.font['S'], text="Block "+str(i)).grid(row=0,column=i,pady=1,padx=15,sticky=E)
                else:
                    Button(self.dataCont, text="Confirm", font=self.font['S'], width=10, bg="#3085d6", command=lambda: self.sendPRASS()).grid(row=maxrow+1,column=i,pady=5,padx=15,sticky=W)

                for j in range(maxrow):
                    exText = str(ws.cell(row=j+1,column=i+1).value)
                    Label(self.dataCont, font=self.font['S'], text=exText).grid(row=j+1,column=i, pady=1, padx=15, sticky=W)
        
    def delete(self,ws,wb,i,root):
        ws.delete_cols(i+1,1)
        wb.save(self.filePath)
        self.root.destroy()
        self.reset(root)

    def sendPRASS(self,):
        self.res = PRASS(self.root,self.inData,self.filePath,self.Wscreen,self.Hscreen).res
        self.root.destroy()
        self.root.quit()
